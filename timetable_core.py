from datetime import datetime, timedelta
import io
try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
except ImportError:
    Workbook = None

DAYS_ORDER = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']


def hex_to_rgba(hex_color, opacity=0.2):
    """Converts hex color to rgba for background styling."""
    hex_color = hex_color.lstrip('#')
    if len(hex_color) == 6:
        r, g, b = tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))
        return f"rgba({r}, {g}, {b}, {opacity})"
    return hex_color


def get_pastel_hex(hex_color):
    """Generates a pastel version of the hex color for Excel background."""
    hex_color = hex_color.lstrip('#')
    if len(hex_color) == 6:
        r, g, b = tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))
        # Blend with white (alpha 0.25)
        alpha = 0.25
        r = int(r * alpha + 255 * (1 - alpha))
        g = int(g * alpha + 255 * (1 - alpha))
        b = int(b * alpha + 255 * (1 - alpha))
        return f"{r:02X}{g:02X}{b:02X}"
    return "E3F2FD"


def process_form_data(form, days_order=None):
    """Takes a Flask `request.form` (or any dict-like with .getlist/.get)
    and returns a dict with schedule data ready for rendering or export.

    Returns keys: schedule, grid_schedule, export_data, breaks, header_times,
    slot_starts, slot_durations, periods_order, days
    """
    days_order = days_order or DAYS_ORDER

    # 1. Extract lists of data from the form
    subjects = form.getlist('subject')
    days = form.getlist('day')
    periods = form.getlist('period')
    durations = form.getlist('duration')
    infos = form.getlist('info')
    colors = form.getlist('color')

    # Format for display (12-hour)
    out_fmt = '%I:%M %p'

    # Get Total Slots
    try:
        total_slots = int(form.get('total_slots', 2))
    except ValueError:
        total_slots = 2

    # Get Time Slot Settings
    default_slots = {
        '1': '08:15', '2': '09:10', '3': '10:05',
        '4': '11:30', '5': '12:25', '6': '13:20'
    }
    slot_starts = {}
    for i in range(1, total_slots + 1):
        key = str(i)
        slot_starts[key] = form.get(f'slot_{key}', default_slots.get(key, '00:00'))

    # Get Slot Durations
    slot_durations = {}
    for i in range(1, total_slots + 1):
        key = str(i)
        try:
            slot_durations[key] = int(form.get(f'duration_{key}', 55))
        except (TypeError, ValueError):
            slot_durations[key] = 55

    # Get Break Settings & Auto-calculate position
    break_names = form.getlist('break_name')
    break_starts = form.getlist('break_start')
    break_ends = form.getlist('break_end')

    breaks = []
    fmt = '%H:%M'
    for i in range(len(break_names)):
        try:
            b_start = datetime.strptime(break_starts[i], fmt)
            b_end = datetime.strptime(break_ends[i], fmt)

            # Auto-calculate 'after' based on time
            after_slot = 0
            for s in range(1, total_slots + 1):
                s_key = str(s)
                s_start = datetime.strptime(slot_starts.get(s_key, '00:00'), fmt)
                s_end_dt = s_start + timedelta(minutes=slot_durations.get(s_key, 55))
                if b_start >= s_end_dt:
                    after_slot = s

            breaks.append({
                'id': f'break_{i}',
                'name': break_names[i],
                'start': b_start.strftime(out_fmt),
                'end': b_end.strftime(out_fmt),
                'after': after_slot
            })
        except (ValueError, IndexError):
            continue

    # Generate periods_order list for consistent iteration
    periods_order = []
    # Insert breaks before any slots (if after=0)
    for b_idx, b in enumerate(breaks):
        if b['after'] == 0:
            periods_order.append(f"break_{b_idx}")
    for i in range(1, total_slots + 1):
        periods_order.append(str(i))
        for b_idx, b in enumerate(breaks):
            if b['after'] == i:
                periods_order.append(f"break_{b_idx}")

    # 2. Structure the data
    raw_schedule = []
    for i in range(len(subjects)):
        if subjects[i].strip():  # Only add if subject name is not empty
            # Calculate times
            start_str = slot_starts.get(periods[i], '00:00')
            fmt = '%H:%M'
            t_start = datetime.strptime(start_str, fmt)

            # Determine duration based on type
            if durations[i] == 'lab3':
                num_slots = 3
            elif durations[i] == 'lab':
                num_slots = 2
            else:
                num_slots = 1

            # Calculate end time based on the last slot covered
            start_p = int(periods[i]) if periods[i].isdigit() else 1
            end_p = start_p + num_slots - 1

            # Ensure end_p is within bounds
            end_p = min(max(end_p, 1), total_slots)

            t_end_start_str = slot_starts.get(str(end_p), '00:00')
            t_end_start = datetime.strptime(t_end_start_str, fmt)
            t_end = t_end_start + timedelta(minutes=slot_durations.get(str(end_p), 55))

            color = colors[i] if i < len(colors) else '#007bff'

            raw_schedule.append({
                'subject': subjects[i],
                'day': days[i],
                'start': t_start.strftime(out_fmt),
                'end': t_end.strftime(out_fmt),
                'info': infos[i],
                'period': periods[i],
                'duration_type': durations[i],
                'color': color,
                'bg_color': hex_to_rgba(color, 0.15),
                'sort_key': t_start
            })

    # 3. Organize data for Grid View (Days vs Time)
    weekly_schedule = {day: [] for day in days_order}
    grid_schedule = {day: {} for day in days_order}

    for item in raw_schedule:
        # Populate list for Export/Legacy use
        if item['day'] in weekly_schedule:
            weekly_schedule[item['day']].append(item)

        # Populate Grid Schedule
        d = item['day']
        p = item['period']
        dtype = item['duration_type']

        # Calculate colspan and mark occupied slots
        colspan = 1
        occupied = []
        start_p = int(p)

        if dtype == 'lab':  # 2 slots
            num_slots = 2
        elif dtype == 'lab3':  # 3 slots
            num_slots = 3
        else:
            num_slots = 1

        end_p = start_p + num_slots - 1

        # Calculate visual colspan including breaks
        visual_colspan = 0
        current_p = start_p
        while current_p <= end_p:
            visual_colspan += 1  # The slot itself
            if current_p > start_p:
                occupied.append(str(current_p))

            # Check for breaks after this slot (if not the last slot)
            if current_p < end_p:
                for b_idx, b in enumerate(breaks):
                    if b['after'] == current_p:
                        visual_colspan += 1
                        occupied.append(f"break_{b_idx}")
            current_p += 1

        item['colspan'] = visual_colspan

        # Check if slot is already taken by a class (dict)
        if p in grid_schedule[d] and isinstance(grid_schedule[d][p], dict):
            # Merge with existing class
            existing = grid_schedule[d][p]
            existing['subject'] += f" + {item['subject']}"
            if item['info']:
                if existing['info']:
                    existing['info'] += f" + {item['info']}"
                else:
                    existing['info'] = item['info']

            # Use the maximum colspan to ensure the merged block covers the widest class
            if item['colspan'] > existing['colspan']:
                existing['colspan'] = item['colspan']
        else:
            grid_schedule[d][p] = item

        for occ in occupied:
            # Only mark if not already a start of a class (dict)
            if occ not in grid_schedule[d] or not isinstance(grid_schedule[d][occ], dict):
                grid_schedule[d][occ] = 'occupied'

    # Detect Gaps (Mark all empty slots as gaps)
    for day in days_order:
        for p in periods_order:
            if p.startswith('break_'):
                continue

            if not grid_schedule[day].get(p):
                grid_schedule[day][p] = {'is_gap': True, 'colspan': 1}

    # 4. Sort classes by start time for each day
    for day in weekly_schedule:
        weekly_schedule[day].sort(key=lambda x: x['sort_key'])

    # Prepare data for export (remove non-serializable objects like datetime)
    export_data = {}
    for day in days_order:
        export_data[day] = []
        # Use grid_schedule to get the final merged items
        for p, item in grid_schedule[day].items():
            if isinstance(item, dict):
                clean_item = {k: v for k, v in item.items() if k != 'sort_key'}
                export_data[day].append(clean_item)

    # Prepare header times for template display
    header_times = {}
    for k, v in slot_starts.items():
        try:
            t_start = datetime.strptime(v, '%H:%M')
            duration = int(slot_durations.get(k, 55))
            t_end = t_start + timedelta(minutes=duration)
            header_times[k] = f"{t_start.strftime(out_fmt)} - {t_end.strftime(out_fmt)}"
        except ValueError:
            header_times[k] = v

    return {
        'schedule': weekly_schedule,
        'grid_schedule': grid_schedule,
        'days': days_order,
        'export_data': export_data,
        'breaks': breaks,
        'header_times': header_times,
        'slot_starts': slot_starts,
        'slot_durations': slot_durations,
        'periods_order': periods_order
    }


def create_excel_bytes(schedule_data, breaks, slot_starts, slot_durations, days_order=None):
    """Create an Excel workbook from schedule data and return BytesIO object."""
    days_order = days_order or DAYS_ORDER

    if Workbook is None:
        raise RuntimeError("openpyxl not installed")

    # Determine total slots from keys
    max_slot = 0
    for k in slot_starts.keys():
        if k.isdigit():
            max_slot = max(max_slot, int(k))
    total_slots = max(max_slot, 2)

    wb = Workbook()
    ws = wb.active
    ws.title = "Weekly Schedule"

    # Define Grid Headers (Day vs Time)
    in_fmt = '%H:%M'
    out_fmt = '%I:%M %p'

    headers = ["Day"]

    for i in range(1, total_slots + 1):
        p = str(i)
        if p in slot_starts:
            t_start = datetime.strptime(slot_starts[p], in_fmt)
            duration = int(slot_durations.get(p, 55))
            t_end = t_start + timedelta(minutes=duration)
            headers.append(f"{t_start.strftime(out_fmt)} - {t_end.strftime(out_fmt)}")
        else:
            headers.append(f"Period {p}")

        # Insert breaks
        for b in breaks:
            if b['after'] == i:
                headers.append(f"{b['start']} - {b['end']}")

    ws.append(headers)

    # Style headers
    header_fill = PatternFill(start_color="343a40", end_color="343a40", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Column Mapping: Period -> Excel Column Index (1-based)
    p_map = {}
    current_col = 2

    for i in range(1, total_slots + 1):
        p_map[str(i)] = current_col
        current_col += 1
        # Skip columns for breaks
        for b in breaks:
            if b['after'] == i:
                current_col += 1

    current_row = 2
    for day in days_order:
        # 1. Write Day Name
        day_cell = ws.cell(row=current_row, column=1, value=day)
        day_cell.font = Font(bold=True)
        day_cell.alignment = Alignment(horizontal="center", vertical="center")

        # 2. Write Break Columns
        col_counter = 2
        for i in range(1, total_slots + 1):
            col_counter += 1
            for b in breaks:
                if b['after'] == i:
                    break_cell = ws.cell(row=current_row, column=col_counter, value=b['name'])
                    break_cell.alignment = Alignment(text_rotation=90, horizontal="center", vertical="center")
                    break_cell.fill = PatternFill(start_color="f0f0f0", end_color="f0f0f0", fill_type="solid")
                    col_counter += 1

        # 3. Fill Classes
        if day in schedule_data:
            for item in schedule_data[day]:
                p = item.get('period')
                if p in p_map:
                    col_idx = p_map[p]
                    colspan = item.get('colspan', 1)

                    # Construct Text
                    text = f"{item['subject']}\n{item['start']} - {item['end']}"
                    if item.get('info'):
                        text += f"\n{item['info']}"

                    cell = ws.cell(row=current_row, column=col_idx, value=text)

                    # Apply Color
                    if item.get('color'):
                        pastel_hex = get_pastel_hex(item['color'])
                        cell.fill = PatternFill(start_color=pastel_hex, end_color=pastel_hex, fill_type="solid")

                    # Merge Cells if needed (e.g. Labs)
                    if colspan > 1:
                        ws.merge_cells(start_row=current_row, start_column=col_idx, end_row=current_row, end_column=col_idx + colspan - 1)

        current_row += 1

    # Apply borders and alignment to all cells
    max_col_idx = current_col - 1
    for row in ws.iter_rows(min_row=1, max_row=current_row - 1, min_col=1, max_col=max_col_idx):
        for cell in row:
            cell.border = thin_border
            is_break_col = cell.value in [b['name'] for b in breaks] if cell.row > 1 else False
            if cell.row > 1 and not is_break_col:
                cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

    # Adjust Column Widths
    ws.column_dimensions['A'].width = 15  # Day
    for i in range(2, max_col_idx + 1):
        col_letter = ws.cell(row=1, column=i).column_letter
        ws.column_dimensions[col_letter].width = 20

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out
